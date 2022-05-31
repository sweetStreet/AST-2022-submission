import collections
import openpyxl
from seutil import IOUtils, BashUtils
from ui.Macros import Macros
import json
from ui.labels.keyword_util import preprocess_dom, get_context_label, get_parent_without_clickable, find_element, column_name_list, dump_xlsx, load_xlsx, remove_non_english_words, \
    get_element_label_heuristic
from collections import OrderedDict
from jsonargparse import CLI
from typing import Union, List
import os
from pathlib import Path

class KeywordGen:

    def __init__(self, project_name: str):
        self.project_name = project_name
        self.xlsx_file_path = f"{Macros.results_dir}/paths/{project_name}-labels.xlsx"

    def label_paths_xlsx_format(
            self,
            crawl_paths_folder: str = "",
        ):
        """
        parse Crawljax (FragGen) results to original xlsx format
        """
        row_content = None
        logging_file = Macros.results_dir / "paths" / f"{self.project_name}-label-paths-xlsx-format.log"
        if logging_file.exists():
            BashUtils.run(f"rm {logging_file}")

        filename = self.project_name + "-labels.xlsx"
        result_path = Macros.results_dir / "paths" / filename
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(
            ["id", "details", "current dom", "context dom", "source state url", "target state url", "state diff url",
            "manual label"])
        # # state diff
        # with IOUtils.cd(f"{Macros.project_dir}/DomTreeDiff"):
        #     BashUtils.run("export JAVA_HOME=/Library/Java/JavaVirtualMachines/jdk-11.0.12.jdk/Contents/Home", expected_return_code=0)
        #     BashUtils.run(f"JAVA_HOME={Macros.JAVA_HOME} mvn clean package", expected_return_code=0)

        if not crawl_paths_folder:
            crawl_paths_file = Macros.results_dir / "frag-gen-updated" / self.project_name / "CrawlPaths.json"
        else:
            if not crawl_paths_folder.endswith("/"):
                crawl_paths_folder = crawl_paths_folder + "/"
            crawl_paths_file = crawl_paths_folder + "CrawlPaths.json"

        Path(Macros.results_dir / "paths").mkdir(parents=True, exist_ok=True)
        IOUtils.dump(logging_file, [f"crawl result: {crawl_paths_file}"], IOUtils.Format.txtList, append=True)

        crawl_paths = IOUtils.load(crawl_paths_file)
        appeared_element = set()
        same_element_and_context = 0
        for crawl_path in crawl_paths:
            for clickable_element in crawl_path:
                try:
                    if type(clickable_element) == str:
                        IOUtils.dump(logging_file, [f"clickable_element is string instead of dict: {clickable_element}"], IOUtils.Format.txtList, append=True)
                        continue
                    unique_id = str(clickable_element["identification"]["value"]) + str(clickable_element["target"]["id"])
                    if unique_id in appeared_element:
                        continue
                    else:
                        appeared_element.add(unique_id)
                        label_dict = self.label_edge_nodes(clickable_element)
                        manual_label = ""
                        # manual_label = id_to_manual_label_dict.get(label_dict["eventable_element_id"], "")
                        # if manual_label == "":
                        #     print("missed", label_dict["eventable_element_id"])
                        if label_dict["cur_dom"] is None:
                            IOUtils.dump(logging_file, [f"current dom cannot be found, {label_dict['eventable_element_id']}"], IOUtils.Format.txtList,
                                         append=True)
                            continue
                        row_content = [label_dict["eventable_element_id"], label_dict["eventable_element_details"], label_dict["cur_dom"],
                                       label_dict["context_dom"], label_dict["source_state_url"], label_dict["target_state_url"],
                                       label_dict["state_diff_url"], manual_label]
                        if label_dict["cur_dom"] == label_dict["context_dom"]:
                            same_element_and_context += 1

                        sheet.append(row_content)
                except Exception as e:
                    IOUtils.dump(logging_file, [f"exception: {e}", f"{row_content}"], IOUtils.Format.txtList, append=True)
                    continue
        IOUtils.dump(logging_file, [f"number of elements that have same context as itself: {same_element_and_context}"], IOUtils.Format.txtList, append=True)
        IOUtils.dump(logging_file, [f"total number of elements: {len(appeared_element)}"], IOUtils.Format.txtList, append=True)
        wb.save(result_path)

    def label_edge_nodes(self, clickable_element: dict):
        res = dict()

        eventable_element_id = str(clickable_element["id"])
        res["eventable_element_id"] = eventable_element_id

        eventable_element_details = clickable_element["element"]
        res["eventable_element_details"] = json.dumps(eventable_element_details)

        state_dom = clickable_element["source"]["dom"]
        cur_dom = find_element(state_dom, clickable_element["identification"]["value"].lower(), return_format="str")
        res["cur_dom"] = cur_dom
        context_dom = get_parent_without_clickable(state_dom, clickable_element["identification"]["value"].lower())
        res["context_dom"] = context_dom

        source_state_id = str(clickable_element["source"]["id"])
        target_state_id = str(clickable_element["target"]["id"])
        source_state_name = self.get_state_name(source_state_id)
        target_state_name = self.get_state_name(target_state_id)
        source_state_url = f"https://github.com/konveyor/tackle-test-generator-ui-level/blob/main/testlabeling/results/frag-gen-updated/{self.project_name}/screenshots/{source_state_name}.png"
        res["source_state_url"] = source_state_url
        target_state_url = f"https://github.com/konveyor/tackle-test-generator-ui-level/blob/main/testlabeling/results/frag-gen-updated/{self.project_name}/screenshots/{target_state_name}.png"
        res["target_state_url"] = target_state_url
        state_diff_url = f"https://github.com/konveyor/tackle-test-generator-ui-level/blob/main/testlabeling/results/states/{self.project_name}/diff_{source_state_id}_{target_state_id}.txt"
        res["state_diff_url"] = state_diff_url

        return res

    def get_state_name(self, state_id: Union[str, int]):
        if state_id == 0 or state_id == "0":
            return "index"
        return "state" + str(state_id)

    # for P in 'addressbook' 'claroline' 'collabtive' 'dimeshift' 'jpetstore' 'mantisbt' 'mrbs' 'pagekit' 'petclinic' 'phoenix' 'ppma' 'retroboard' 'splittypie'; do python -m ui.labels.keyword_fraggen --project_name=$P label_with_preprocess --algorithm="keybert"; done
    def label_with_preprocess(self, algorithm: str):
        os.environ["TOKENIZERS_PARALLELISM"] = "false"
        logging_file = Macros.results_dir / algorithm / f"{self.project_name}-labels.log"
        if logging_file.exists():
            IOUtils.rm(logging_file)

        result = []
        col_name_list = column_name_list(self.xlsx_file_path)
        col_name_list.append(f"{algorithm} label")
        result.append(col_name_list)
        index_of_manual_label = col_name_list.index("manual label")
        index_of_context = col_name_list.index("context dom")
        index_of_current_dom = col_name_list.index("current dom")
        index_of_details = col_name_list.index("details")
        num_of_none_label = 0
        for row in load_xlsx(self.xlsx_file_path):
            if row[index_of_manual_label].strip() == "none":
                num_of_none_label += 1
                continue
            if algorithm == "preprocess":
                preprocess_edge_label = preprocess_dom(remove_non_english_words(preprocess_dom(row[index_of_current_dom]))).strip()
                if preprocess_edge_label == "":
                    preprocess_edge_label = preprocess_dom(remove_non_english_words(preprocess_dom(row[index_of_context]))).strip()
                row.append(preprocess_edge_label)
            elif algorithm == "keybert":
                keybert_edge_label = get_context_label(row[index_of_current_dom])
                if keybert_edge_label == "":
                    keybert_edge_label = get_context_label(row[index_of_context])
                if keybert_edge_label == "":
                    IOUtils.dump(logging_file, ["fail to get the label", row[index_of_current_dom]],
                                 IOUtils.Format.txtList,
                                 append=True)
                row.append(keybert_edge_label)
            elif algorithm == "heuristic":
                (Macros.results_dir / algorithm).mkdir(exist_ok=True)
                heuristic_edge_label = get_element_label_heuristic(json.loads(row[index_of_details]))
                if heuristic_edge_label.strip() == "":
                    heuristic_edge_label = get_context_label(row[index_of_context])
                row.append(heuristic_edge_label)
            result.append(row)
        IOUtils.dump(logging_file, [f"number of labels that are not visible: {num_of_none_label}"],
                     IOUtils.Format.txtList,
                     append=True)
        dump_xlsx(Macros.results_dir/algorithm/f"{self.project_name}-labels.xlsx", result)

    def data_count_attributes_per_project(self, projects: List[str]):
        '''
        :param projects: name of the projects to rank the attributes
        '''
        data = collections.defaultdict(list)
        keyset = set()
        for project in projects:
            data_source_counter = IOUtils.load(f"{Macros.results_dir}/metrics/{project}-label-source.json")
            # ordered_data_source_counter = collections.OrderedDict(sorted(data_source_counter.items()))
            ordered_data_source_counter = collections.OrderedDict(
                sorted(data_source_counter.items(), key=lambda item: item[1], reverse=True))
            data[project] = ordered_data_source_counter
            keyset.update(ordered_data_source_counter.keys())
        large_dict = collections.Counter()
        for project in projects:
            for key in keyset:
                if key not in data[project]:
                    data[project][key] = 0
                else:
                    large_dict[key] += data[project][key]
        sorted_keys = list(dict(sorted(large_dict.items(), key=lambda item: item[1], reverse=True)).keys())
        reordered_data = collections.OrderedDict()
        for project in projects:
            reordered_data[project] = dict()
            for key in sorted_keys:
                if key in data[project]:
                    reordered_data[project][key] = data[project][key]
                else:
                    reordered_data[project][key] = 0
        return reordered_data, list(sorted_keys)

    def xlsx_to_json(self, crawl_paths_folder: str = "", method: str = "PCFG"):
        logging_file = Macros.results_dir / "paths" / f"{self.project_name}-labels-offline-log.log"
        if logging_file.exists():
            BashUtils.run(f"rm {logging_file}")

        id_to_label = dict()
        excel_path = Macros.results_dir / "paths" / f"{self.project_name}-labels.xlsx"
        wb_obj = openpyxl.load_workbook(excel_path)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        for i in range(1, m_row + 1):
            try:
                labels_dict = {}
                label_id_obj = sheet_obj.cell(row=i, column=1)
                label_id = label_id_obj.value
                where_to_get_obj = sheet_obj.cell(row=i, column=11)
                if where_to_get_obj.value == "none":
                    labels_dict["label"] = "none"
                    labels_dict["edge_label"] = "none"
                    labels_dict["context_label"] = "none"
                    labels_dict["diff_label"] = "none"
                    labels_dict["manual_label"] = "none"
                    id_to_label[label_id] = labels_dict
                    continue
                edge_label_obj = sheet_obj.cell(row=i, column=8)
                labels_dict["edge_label"] = edge_label_obj.value.strip()
                diff_label_obj = sheet_obj.cell(row=i, column=9)
                if diff_label_obj.value:
                    labels_dict["diff_label"] = diff_label_obj.value.split(",")[0].replace("'", "").replace("[", "").replace("(", "").strip()
                else:
                    labels_dict["diff_label"] = ""

                context_label_obj = sheet_obj.cell(row=i, column=10)
                if context_label_obj.value:
                    labels_dict["context_label"] = labels_dict["context_label"] = context_label_obj.value.split(",")[0]\
                        .replace("'", "").replace("[", "").replace("(", "").strip()
                else:
                    labels_dict["context_label"] = ""

                manual_label_obj = sheet_obj.cell(row=i, column=12)
                labels_dict["manual_label"] = manual_label_obj.value

                word_set = set()
                if labels_dict["edge_label"]:
                    word_set.update(labels_dict["edge_label"].split(" "))
                if labels_dict["diff_label"]:
                    word_set.update(labels_dict["diff_label"].split(" "))
                if labels_dict["context_label"]:
                    word_set.update(labels_dict["context_label"].split(" "))
                labels_dict["label"] = " ".join(word_set).strip()
                id_to_label[label_id] = labels_dict
            except Exception as e:
                IOUtils.dump(logging_file, [f"row {i}: {e}"], IOUtils.Format.txtList, append=True)
                continue

        if not crawl_paths_folder:
            crawl_paths_file = Macros.results_dir / "frag-gen-updated" / self.project_name / "CrawlPaths.json"
        else:
            if not crawl_paths_folder.endswith("/"):
                crawl_paths_folder = crawl_paths_folder + "/"
            crawl_paths_file = crawl_paths_folder + "CrawlPaths.json"

        crawl_paths = IOUtils.load(crawl_paths_file)

        json_result_path = Macros.results_dir / "paths" / f"{self.project_name}-labels.json"
        result = []

        appeared_element_to_id = dict()
        for crawl_path in crawl_paths:
            method_name = ""
            test_case = OrderedDict()
            for clickable_element in crawl_path:
                try:
                    clickable_id = str(clickable_element["id"])
                    method_name += clickable_id + "_"
                    unique_id = str(clickable_element["identification"]["value"]) + str(clickable_element["target"]["id"])
                    if clickable_id in id_to_label:
                        label_dict = id_to_label[clickable_id]
                        appeared_element_to_id[unique_id] = clickable_id
                    else:
                        if unique_id in appeared_element_to_id and appeared_element_to_id[unique_id] in id_to_label:
                            label_dict = id_to_label[appeared_element_to_id[unique_id]]
                        else:
                            label_dict = {"label": "none", "edge_label": "none", "context_label": "none",
                                          "diff_label": "none", "manual_label": "none"}

                    label_dict["details"] = clickable_element["element"]
                    test_case[clickable_id] = label_dict
                except Exception as e:
                    IOUtils.dump(logging_file, [f"{e} : {clickable_element}"], IOUtils.Format.txtList, append=True)
                    continue
            result.append({"label_map": test_case, "method_name": method_name[:len(method_name)-1]})
        IOUtils.dump(json_result_path, result, IOUtils.Format.jsonNoSort)

    def check_label_source_helper(self, dom: dict, manual_label: str):
        result = []
        for key, value in dom.items():
            preprocessed = preprocess_dom(value)
            for word in preprocessed.split(" "):
                if word.strip() == "":
                    continue
                if word in manual_label:
                    result.append(key)
                    break
        return result

    def check_label_source(self):
        xlsx_file = Macros.results_dir / "paths" / f"{self.project_name}-labels.xlsx"
        col_name_list = column_name_list(xlsx_file)

        index_of_manual_label = col_name_list.index("manual label")
        index_of_details = col_name_list.index("details")

        label_source_counter = collections.Counter()
        for row in load_xlsx(xlsx_file):
            if row[index_of_manual_label] is None:
                print(f"{xlsx_file} at line {index_of_manual_label} manual label is empty")
                return
            if row[index_of_manual_label].strip() == "none":
                continue
            dom_dict = json.loads(row[index_of_details]).get("attributes")
            dom_dict["text"] = json.loads(row[index_of_details]).get("text")
            label_sources = self.check_label_source_helper(dom_dict, row[index_of_manual_label])
            for label_source in label_sources:
                # exclude these two attributes because they are generated by Crawljax
                if label_source in ["evlist", "anyaccess"]:
                    continue
                label_source_counter[label_source] += 1
        IOUtils.dump(f"{Macros.results_dir}/metrics/{self.project_name}-label-source.json", label_source_counter, IOUtils.Format.json)

if __name__ == "__main__":
    CLI(KeywordGen, as_positional=False)
