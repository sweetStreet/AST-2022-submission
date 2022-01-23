import json
import string
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import collections
from gensim.models import Phrases
from gensim.models.phrases import Phraser
from seutil import IOUtils, BashUtils
from ui.Macros import Macros
from bs4 import BeautifulSoup
import tomotopy as tp
import enchant
from typing import List
from keybert import KeyBERT
import csv
from ui.labels import compare_dom_tree
import os
import time
from pathlib import Path
from lxml import html, etree
from io import StringIO
from summa.summarizer import summarize
from summa.keywords import keywords
import xmltodict as xd
import re
import logging
from sentence_transformers import SentenceTransformer
import openpyxl
from typing import Union

'''
Code for generating labels
'''
# import nltk
# nltk.download('stopwords')
# import nltk
# nltk.download('wordnet')

html_stop_words = ["a", "abbr", "acronym", "address", "area", "b", "base", "bdo", "big", "blockquote", "body", "br",
              "button", "caption", "cite", "code", "col", "colgroup", "dd", "del", "dfn", "div", "dl", "doctype",
              "dt", "em", "fieldset", "form", "h1", "h2", "h3", "h4", "h5", "h6", "head", "html", "hr", "i", "img",
              "input", "ins", "kbd", "label", "legend", "li", "link", "map", "meta", "noscript", "object", "ol",
              "optgroup", "option", "p", "param", "pre", "q", "samp", "script", "select", "small", "span", "strong",
              "style", "sub", "sup", "table", "tbody", "td", "text", "textarea", "tfoot", "th", "thead", "title",
              "tr", "tt", "ul", "var", "", "submit", "false", "action", "javascript", "javascript:;", "href"]

########################################## NLP preprocessing ########################################################
def preprocess_dom(state_stripped_dom: str):
    # print(state_stripped_dom)
    # tokenize, including split words with camel case, to lower case, removing punctuation, removing letter
    tokenized_dom = tokenize(state_stripped_dom)
    # remove stop words
    remove_stop_words_dom = remove_stop_words(tokenized_dom)
    # lemmatization
    lemmatized_dom = lemmatization(remove_stop_words_dom)
    # remove stop words again
    remove_stop_words_dom = remove_stop_words(lemmatized_dom)
    return remove_stop_words_dom


def tokenize(s: str):
    result = ""
    buffer = ""
    for word in s.split():
        if word.isupper():
            if len(word) > 1:
                result += word.lower() + " "
        else:
            for c in word:
                if c in string.ascii_lowercase:
                    buffer += c
                elif c in string.ascii_uppercase:
                    if buffer != "":
                        if len(buffer) > 1:
                            result += buffer + " "
                        buffer = ""
                    buffer += c.lower()
                else:
                    if buffer != "":
                        if len(buffer) > 1:
                            result += buffer + " "
                        buffer = ""
            if buffer != "":
                if len(buffer) > 1:
                    result += buffer + " "
                buffer = ""
    return result


def remove_stop_words(s: str):
    english_stop_words = stopwords.words('english')
    # stop_words.extend(
    #     ['<html><head><meta', 'http', 'id', 'src', 'images', 'content', 'type', 'text', 'html', 'gif', 'txt', "alt",
    #      "www"])
    # html tags
    dictionary = enchant.Dict("en_US")
    # "separator", "coords", "sm", "nbsp"
    # remove stop words and non-english words
    # enchant package requires C library https://pyenchant.github.io/pyenchant/install.html
    remove_stop_words_dom = ""
    for token in s.split(" "):
        # or token in english_stop_words
        if token in html_stop_words or not dictionary.check(token):
            continue
        else:
            remove_stop_words_dom = remove_stop_words_dom + " " + token
    return remove_stop_words_dom


def lemmatization(s: str):
    lemmatizer = WordNetLemmatizer()
    lemmatization_dom = ""
    for token in s.split(" "):
        lemmatization_dom = lemmatization_dom + " " + lemmatizer.lemmatize(token)
    return lemmatization_dom


def remove_non_english_words(s: str):
    dictionary = enchant.Dict("en_US")
    remove_non_english_words_dom = ""
    for token in s.split(" "):
        # or token in english_stop_words
        if not token or not dictionary.check(token):
            continue
        else:
            remove_non_english_words_dom = remove_non_english_words_dom + " " + token
    return remove_non_english_words_dom


def unigram(tokens):
    model = collections.defaultdict(lambda: 0.01)
    for f in tokens:
        try:
            model[f] += 1
        except KeyError:
            model[f] = 1
            continue
    N = float(sum(model.values()))
    for word in model:
        model[word] = model[word] / N
    return model


########################################## label edge (clickable element) ############################################
def process_attribute_value(text):
    element_label = ""
    dictionary = enchant.Dict("en_US")
    punc = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    text = remove_non_english_words(remove_stop_words(lemmatization(tokenize(text))))
    english = False
    if text:
        for button_text_word in text.split(" "):
            if button_text_word in punc:
                continue
            if not dictionary.check(button_text_word.lower()):
                english = False
                break
            english = True
        if english:
            element_label = text
    logging.debug(f"text: {text}, isEnglish: {english}, element label: {element_label}")
    return element_label.strip()

def parse_dom_to_dict(dom: str):
    context_dom_str = r"<html>" + dom + r"</html>"
    context_dom_dict = xd.parse(context_dom_str)
    logging.debug(f"context dom dict: {context_dom_dict}")
    return context_dom_dict

def get_context_label(context_dom: str):
    text = dict_value_to_str(parse_dom_to_dict(context_dom))
    preprocessed = remove_non_english_words(lemmatization(tokenize(preprocess_dom(text))))
    kw_model = KeyBERT()
    keywords = kw_model.extract_keywords(preprocessed, keyphrase_ngram_range=(1, 2), stop_words=html_stop_words, use_mmr=True,
                                         diversity=0.8)
    print(keywords)
    logging.debug(f"keywords: {keywords}")
    if keywords:
        element_label = str(keywords[0][0])
        # element_label = str(keywords)
        logging.debug(f"keyword: {element_label}")
    else:
        element_label_list = []
        for word in preprocessed.split(" "):
            if len(word.strip()) > 0 and word.strip() not in element_label_list:
                element_label_list.append(word.strip())
        element_label = " ".join(element_label_list)
    element_label = process_attribute_value(element_label)
    return element_label.strip()

def dict_value_to_str(context_dom_input):
    if type(context_dom_input) == dict or type(context_dom_input) == collections.OrderedDict:
        res = ""
        for value in context_dom_input.values():
            res += dict_value_to_str(value)
        return res
    elif type(context_dom_input) == str:
        return context_dom_input
    elif type(context_dom_input) == list:
        res = ""
        for context_dom_item in context_dom_input:
            res += dict_value_to_str(context_dom_item)
        return res
    elif type(context_dom_input) == tuple:
        if len(context_dom_input) == 2:
            return context_dom_input[1]
        else:
            return " ".join(context_dom_input)
    else:
        logging.debug(f"missed {type(context_dom_input)}")
        return ""

def dom_dict_value_to_list(context_dom_input: Union[dict, str, list]):
    if isinstance(context_dom_input, str):
        preprocessed = preprocess_dom(context_dom_input)
        if preprocessed:
            return [preprocessed]
        else:
            return []
    elif isinstance(context_dom_input, dict) or isinstance(context_dom_input, collections.OrderedDict):
        res = []
        for value in context_dom_input.values():
            for r in dom_dict_value_to_list(value):
                res.append(r)
        return res
    elif isinstance(context_dom_input, list):
        res = []
        for context_dom_item in context_dom_input:
            for r in dom_dict_value_to_list(context_dom_item):
                res.append(r)
        return res
    else:
        return []

def get_parent_without_clickable(state_dom: str, eventable_element_details: str):
    cur_dom = find_element(state_dom, eventable_element_details)
    if cur_dom is None:
        return " "
    # stop when find more than on clickable element
    # maximum number of iterations is 3, which means that at most find the forth parent
    iterations = 0
    while cur_dom.getparent() and iterations <= 3:
        try:
            cur_dom_string = etree.tostring(cur_dom).decode('UTF-8').strip()
            if not cur_dom_string.endswith(">"):
                cur_dom_string = cur_dom_string[:cur_dom_string.rindex(">") + 1]
            cur_dom_without_context = etree.parse(StringIO(cur_dom_string))
        except Exception as e:
            print(cur_dom_string)
            print("get parent", e)
        parent = cur_dom.getparent()
        if (parent is None) or (parent and contain_more_than_one_clickable(cur_dom_without_context)):
            return etree.tostring(cur_dom).decode('UTF-8')
        cur_dom = parent
        iterations += 1
    return etree.tostring(cur_dom).decode('UTF-8')


def contain_more_than_one_clickable(dom):
    # <a> or <input> with type "submit" (//input[@type='submit']) or <button>
    # <span> with framework wiget such as "data-react-toolbox"
    # print("dom", etree.tostring(dom).decode('UTF-8'))
    xpath_res = dom.xpath("//*[self::a or self::input[@type='submit'] or self::button]")
    logging.debug(f"xpath result: {xpath_res}")
    logging.debug(f"length: {len(xpath_res)}")
    if len(xpath_res) > 1:
        return True
    return False


def preprocess_eventable_element_details(eventable_element_details):
    text = ""
    for value in eventable_element_details.values():
        if type(value) == dict:
            text = text + json.dumps(list(value.values())) + " "
        elif type(value) == str:
            text = text + value + " "
    logging.debug(f"text: {text}")
    preprocessed = preprocess_dom(text)
    return preprocessed


def get_element_label_heuristic(eventable_element_details: dict):
    # heuristic rule
    element_label = ""
    if element_label == "" and "text" in eventable_element_details:
        button_text = eventable_element_details["text"]
        element_label = process_attribute_value(button_text)
    if element_label == "" and "attributes" in eventable_element_details and "href" in eventable_element_details[
        "attributes"]:
        href = eventable_element_details["attributes"]["href"]
        tokenized_href = process_attribute_value(href.strip().split("?")[-1])
        kw_model = KeyBERT()
        keywords = kw_model.extract_keywords(tokenized_href, keyphrase_ngram_range=(1, 2), stop_words='english', use_mmr=True,
                                             diversity=0.7)
        if keywords:
            element_label = keywords[0][0]
        else:
            element_label = href.strip().split("/")[-1].split(".")[0]
            element_label = process_attribute_value(element_label)
    if element_label == "" and "attributes" in eventable_element_details and "class" in eventable_element_details[
            "attributes"]:
        value_text = eventable_element_details["attributes"]["class"]
        element_label = process_attribute_value(value_text)
    if element_label == "" and "attributes" in eventable_element_details and "title" in eventable_element_details[
        "attributes"]:
        value_text = eventable_element_details["attributes"]["title"]
        element_label = process_attribute_value(value_text)
    if element_label == "" and "attributes" in eventable_element_details and "aria-label" in eventable_element_details[
            "attributes"]:
        value_text = eventable_element_details["attributes"]["aria-label"]
        element_label = process_attribute_value(value_text)
    if element_label == "" and "attributes" in eventable_element_details and "id" in eventable_element_details[
        "attributes"]:
        name_text = eventable_element_details["attributes"]["id"]
        element_label = process_attribute_value(name_text)
    if element_label == "" and "attributes" in eventable_element_details and "value" in eventable_element_details[
        "attributes"]:
        name_text = eventable_element_details["attributes"]["value"]
        element_label = process_attribute_value(name_text)
    if element_label == "":
        preprocessed = remove_non_english_words(
            lemmatization(tokenize(preprocess_eventable_element_details(eventable_element_details))))
        kw_model = KeyBERT()
        keywords = kw_model.extract_keywords(preprocessed, keyphrase_ngram_range=(1, 2), stop_words='english',
                                             use_mmr=True, diversity=0.7)
        if keywords:
            element_label = keywords[0][0]
        else:
            element_label = preprocessed
    # pos_tag([word])[0][1].startswith('V')
    # print("element label", element_label)
    return element_label

def get_element_label(eventable_element_details: dict):
    # instead of applying heuristic rule, directly parse all the attributes to KeyBert
    # preprocessed = preprocess_eventable_element_details(eventable_element_details)
    # logging.debug(f"preprocessed: {preprocessed}")
    preprocessed = remove_non_english_words(lemmatization(tokenize(preprocess_eventable_element_details(eventable_element_details))))
    kw_model = KeyBERT()
    keywords = kw_model.extract_keywords(preprocessed, keyphrase_ngram_range=(1, 2), stop_words=html_stop_words, use_mmr=True,
                                         diversity=0.8)
    logging.debug(f"keywords: {keywords}")
    if len(keywords) > 0:
        element_label = str(keywords[0][0])
        logging.debug(f"keyword: {element_label}")
    else:
        element_label = preprocessed
    return element_label


########################################## label node (state) ############################################
def find_parent(project_name: str):
    # CrawlPaths element contains too many newly added attributes
    crawl_paths_file = Macros.results_dir / "frag-gen" / project_name / "CrawlPaths.json"
    crawl_paths = IOUtils.load(crawl_paths_file)

    # open the file in the write mode
    with open(f'{Macros.results_dir}/{project_name}-dom.csv', 'w') as f:
        # create the csv writer
        writer = csv.writer(f)
        writer.writerow(["id", "element", "first parent", "second parent", "third parent", "forth parent"])

        for crawl_path in crawl_paths:
            for clickable_element in crawl_path:
                previous_state = clickable_element["source"]
                previous_state_id = str(clickable_element["source"]["id"])
                if previous_state_id == "0":
                    html_path = f'{Macros.results_dir}/frag-gen/{project_name}/doms/index.html'
                else:
                    html_path = f'{Macros.results_dir}/frag-gen/{project_name}/doms/state{previous_state_id}.html'
                with open(html_path, 'r') as file:
                    state_dom = file.read()
                if not state_dom:
                    logging.debug("failed to find the state dom")
                    continue
                eventable_element_xpath = str(clickable_element["identification"]["value"])
                cur_dom, first_parent, second_parent, third_parent, forth_parent = find_parent_for_each_element(
                    state_dom, eventable_element_xpath.lower())
                if cur_dom == "":
                    logging.debug("Could not find element")
                    logging.debug(f'id: {clickable_element["id"]}')
                    logging.debug(f'element: {clickable_element["element"]}')
                    logging.debug(f'state id: {previous_state["id"]}')
                    continue
                # write a row to the csv file
                writer.writerow(
                    [clickable_element["id"], cur_dom, first_parent, second_parent, third_parent, forth_parent])


def find_element(state_dom: str, clickable_element_xpath: str, return_format: str = "dom"):
    '''
    locate the clickable element in the web page
    :param state_dom: string format of html
    :param clickable_element_xpath: the xpath of the element, make sure that xpath is in lowercase
    :return: HtmlElement
    '''
    try:
        tree = html.fromstring(state_dom)
        element = tree.xpath(clickable_element_xpath)
        if not element or len(element) == 0:
            return None
        if return_format == "str":
            return etree.tostring(element[0]).decode('UTF-8')
        return element[0]
    except Exception as e:
        logging.debug(clickable_element_xpath)
        logging.debug(e)
        return None

########################################## openpyxl operations ############################################
def column_name_list(xlsx_file_path: str):
    """
    get the column index with the column name
    row is a tuple of cell object from openpyxl
    name is column name, the type is string
    """
    col_name_list = []
    wb_obj = openpyxl.load_workbook(xlsx_file_path)
    sheet_obj = wb_obj.active
    for index, row in enumerate(sheet_obj.iter_rows()):
        if index == 0:
            for cell in row:
                col_name_list.append(cell.value)
            break
    return col_name_list

def dump_xlsx(xlsx_file_path: str, content: List[List[str]]):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in content:
        ws.append(row)
    # Save the file
    wb.save(xlsx_file_path)

def load_xlsx(xlsx_path: str):
    """
    read xlsx file
    the return type is a list of string
    """
    wb_obj = openpyxl.load_workbook(xlsx_path)
    sheet_obj = wb_obj.active
    for index, row in enumerate(sheet_obj.iter_rows()):
        if index == 0:
            continue
        yield [cell.value for cell in row]