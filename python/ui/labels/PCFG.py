from typing import Union, List

from nltk import Tree, ParentedTree
from nltk.corpus import treebank
from nltk import Nonterminal
from nltk import induce_pcfg
from nltk.parse import pchart, ViterbiParser
from ui.labels.keyword_util import find_element, tokenize, column_name_list, load_xlsx, preprocess_dom, lemmatization, \
    remove_non_english_words, parse_dom_to_dict, dom_dict_value_to_list, get_context_label
from ui.Macros import Macros
import json
import openpyxl
import collections
from stanza.server import CoreNLPClient
from seutil import IOUtils
from jsonargparse import CLI
from nltk.corpus import verbnet

class PCFG:
    def __init__(self, project_name: str):
        self.project_name = project_name
        self.logging_file = Macros.results_dir / "pcfg" / f"process-pcfg-labels-{project_name}.log"

    # for i in 'addressbook' 'claroline' 'collabtive' 'jpetstore' 'mantisbt' 'mrbs' 'phoenix' 'ppma';
    # do python -m ui.labels.PCFG --project_name $i apply_pcfg_postag; done;
    def apply_pcfg_postag(self):
        if self.logging_file.exists():
            IOUtils.rm(self.logging_file)

        xlsx_path = f"{Macros.results_dir}/paths/{self.project_name}-labels.xlsx"

        num_elements_not_visible = 0
        element_details_to_pcfg_index = dict()
        pcfg_index_to_verb_keyword = collections.defaultdict(tuple)
        pcfg_index_to_noun_keyword = collections.defaultdict(tuple)
        pcfg_index_to_tree = collections.defaultdict(str)
        html_list = []
        productions = []

        col_name_list = column_name_list(xlsx_path)
        index_of_manual_label = col_name_list.index("manual label")
        index_of_element_details = col_name_list.index("details")
        index_of_current_dom = col_name_list.index("current dom")
        index_of_context_dom = col_name_list.index("context dom")

        for index, row in enumerate(load_xlsx(xlsx_path)):
            # ignore those that are not visible
            if row[index_of_manual_label] and row[index_of_manual_label] == "none":
                num_elements_not_visible += 1
                continue
            # details of clickable
            if row[index_of_element_details] in element_details_to_pcfg_index.keys():
                continue
            element_details_to_pcfg_index[row[index_of_element_details]] = len(element_details_to_pcfg_index.keys())
            splitted_html = dom_dict_value_to_list(parse_dom_to_dict(row[index_of_current_dom]))
            if len(splitted_html) > 0:
                html_list.append(" ".join(splitted_html))

        pos_tagged_html_list = self.pos_tag(html_list)
        for str_tree in pos_tagged_html_list:
            tree = Tree.fromstring(str_tree)
            productions += tree.productions()
        S = Nonterminal('a')
        grammar = induce_pcfg(S, productions)

        production_rule_dict = collections.defaultdict(list)
        for line_index, grammar_line in enumerate(str(grammar).split("\n")):
            if line_index == 0:
                continue
            try:
                parent = grammar_line.split("->")[0].strip()
                child = grammar_line.split("->")[1].split(r"[")[0].strip()
                possibility = float(grammar_line.split("->")[1].split(r"[")[-1].replace("]", ""))
                production_rule_dict[parent].append((child, possibility))
                IOUtils.dump(self.logging_file, [f"{parent}, {child}, {possibility} \n"], IOUtils.Format.txtList, append=True)
            except Exception as e:
                IOUtils.dump(self.logging_file, [e, grammar_line], IOUtils.Format.txtList, append=True)
                raise RuntimeError(e)

        for index, str_tree in enumerate(pos_tagged_html_list):
            parented_tree = ParentedTree.fromstring(str_tree)
            np_trees = self.traverse(parented_tree, production_rule_dict, "V")
            IOUtils.dump(self.logging_file, [f"verb trees: {np_trees}"], IOUtils.Format.txtList, append=True)
            noun_np_trees = self.traverse(parented_tree, production_rule_dict, "N")
            if np_trees:
                try:
                    sorted_np_trees = sorted(np_trees, key=lambda x: x[1][0], reverse=True)
                    pcfg_label = " ".join(sorted_np_trees[0][0].leaves())
                    pcfg_index_to_verb_keyword[index] = (pcfg_label, sorted_np_trees[0][1][0], sorted_np_trees[0][1][1])
                    pcfg_index_to_tree[index] = str_tree
                except TypeError as e:
                    pcfg_index_to_tree[index] = str_tree
                    IOUtils.dump(self.logging_file, [f"{e}"], IOUtils.Format.txtList, append=True)
                    IOUtils.dump(self.logging_file, [t for t in np_trees], IOUtils.Format.txtList, append=True)
                    raise RuntimeError(e)
            else:
                pcfg_label = ""
                pcfg_index_to_verb_keyword[index] = (pcfg_label, "")
                pcfg_index_to_tree[index] = str_tree

            if noun_np_trees:
                try:
                    sorted_noun_np_trees = sorted(noun_np_trees, key=lambda x: x[1][0], reverse=True)
                    pcfg_noun_label = " ".join(sorted_noun_np_trees[0][0].leaves())
                    # (label, type, possibility)
                    pcfg_index_to_noun_keyword[index] = (pcfg_noun_label, sorted_noun_np_trees[0][1][0], sorted_noun_np_trees[0][1][1])
                except Exception as e:
                    print(e)
                    raise RuntimeError(e)
            else:
                pcfg_noun_label = ""
                pcfg_index_to_noun_keyword[index] = (pcfg_noun_label, "")

        pcfg_wb_obj = openpyxl.Workbook()
        pcfg_sheet = pcfg_wb_obj.active
        pcfg_row = list(col_name_list)
        pcfg_row.append("pcfg label")
        pcfg_row.append("tree")
        pcfg_sheet.append(pcfg_row)
        for index, row in enumerate(load_xlsx(xlsx_path)):
            IOUtils.dump(self.logging_file, [f"element: {index}"], IOUtils.Format.txtList, append=True)
            if row[index_of_manual_label] and row[index_of_manual_label] == "none":
                num_elements_not_visible += 1
                continue

            pcfg_index = element_details_to_pcfg_index[row[index_of_element_details]]
            if pcfg_index_to_verb_keyword[pcfg_index] and pcfg_index_to_verb_keyword[pcfg_index][0] and len(pcfg_index_to_verb_keyword[pcfg_index][0].split(" ")) < 5:
                pcfg_label = pcfg_index_to_verb_keyword[pcfg_index][0]
                IOUtils.dump(self.logging_file, [f"verb(pcfg): {pcfg_index_to_verb_keyword[pcfg_index]}"], IOUtils.Format.txtList, append=True)
            else:
                pcfg_label = self.heuristic_rule_verb(row[index_of_current_dom], row[index_of_context_dom])
                IOUtils.dump(self.logging_file, [f"current dom: {row[index_of_current_dom]}"],
                             IOUtils.Format.txtList, append=True)
                IOUtils.dump(self.logging_file, [f"verb(verbnet): {self.heuristic_rule_verb(row[index_of_current_dom], row[index_of_context_dom])}"],
                             IOUtils.Format.txtList, append=True)
            if pcfg_index_to_verb_keyword[pcfg_index] and pcfg_index_to_verb_keyword[pcfg_index][0] and len(pcfg_index_to_verb_keyword[pcfg_index][0].split(" ")) < 5 and "VP" not in pcfg_index_to_verb_keyword[pcfg_index][1] \
                    and pcfg_index_to_noun_keyword[pcfg_index] and len(pcfg_index_to_noun_keyword[pcfg_index][0].split(" ")) < 5:
                pcfg_label += " " + pcfg_index_to_noun_keyword[pcfg_index][0]
                IOUtils.dump(self.logging_file, [f"noun: {pcfg_index_to_noun_keyword[pcfg_index]}"], IOUtils.Format.txtList,
                         append=True)
            postag_tree = pcfg_index_to_tree[pcfg_index]
            IOUtils.dump(self.logging_file, [f"tree: {postag_tree} \n"], IOUtils.Format.txtList, append=True)

            pcfg_row = list(row)
            remove_duplicate_word_pcfg_label = ""
            for word in pcfg_label.split(" ")[::-1]:
                if not word:
                    continue
                if word not in remove_duplicate_word_pcfg_label:
                    remove_duplicate_word_pcfg_label = word + " " + remove_duplicate_word_pcfg_label
            pcfg_row.append(remove_duplicate_word_pcfg_label.strip())
            pcfg_row.append(postag_tree)
            pcfg_sheet.append(pcfg_row)

        pcfg_xlsx_path = f"{Macros.results_dir}/pcfg/{self.project_name}-pcfg-labels2.xlsx"
        pcfg_wb_obj.save(pcfg_xlsx_path)

    # for i in 'addressbook' 'claroline' 'collabtive' 'jpetstore' 'mantisbt' 'mrbs' 'phoenix' 'ppma';
    # do python -m ui.labels.PCFG --project_name $i apply_pcfg_postag_forest; done;
    def apply_pcfg_postag_forest(self):
        if self.logging_file.exists():
            IOUtils.rm(self.logging_file)

        num_elements_not_visible = 0
        productions = []
        xlsx_path = f"{Macros.results_dir}/paths/{self.project_name}-labels.xlsx"
        html_list = []
        each_html_len = []
        html_list_index_to_sheet_index = dict()
        element_details_to_sheet_index = dict()
        col_name_list = column_name_list(xlsx_path)
        index_of_manual_label = col_name_list.index("manual label")
        index_of_element_details = col_name_list.index("details")
        index_of_current_dom = col_name_list.index("current dom")
        index_of_context_dom = col_name_list.index("context dom")
        index_of_state_diff = col_name_list.index("state diff url")
        index_of_id = col_name_list.index("id")
        for index, row in enumerate(load_xlsx(xlsx_path)):
            # ignore those that are not visible
            if row[index_of_manual_label] and row[index_of_manual_label] == "none":
                num_elements_not_visible += 1
                continue
            # details of clickable
            if row[index_of_element_details] in element_details_to_sheet_index.keys():
                continue
            element_details_to_sheet_index[row[index_of_element_details]] = index

            splitted_html = dom_dict_value_to_list(parse_dom_to_dict(row[index_of_current_dom]))
            if len(splitted_html) > 0:
                html_list.extend(splitted_html)
            else:
                splitted_html = dom_dict_value_to_list(parse_dom_to_dict(row[index_of_context_dom]))
                if len(splitted_html) > 0:
                    html_list.extend(splitted_html)
                else:
                    raise RuntimeError("fail to get any useful string")

            if each_html_len:
                each_html_len.append(each_html_len[-1] + len(splitted_html))
            else:
                each_html_len.append(len(splitted_html))
            html_list_index_to_sheet_index[len(each_html_len) - 1] = index
        pos_tagged_html_list = self.pos_tag(html_list)
        for str_tree in pos_tagged_html_list:
            tree = Tree.fromstring(str_tree)
            productions += tree.productions()
        S = Nonterminal('a')
        grammar = induce_pcfg(S, productions)

        production_rule_dict = collections.defaultdict(list)
        for line_index, grammar_line in enumerate(str(grammar).split("\n")):
            if line_index == 0:
                continue
            try:
                parent = grammar_line.split("->")[0].strip()
                child = grammar_line.split("->")[1].split(r"[")[0].strip()
                possibility = float(grammar_line.split("->")[1].split(r"[")[-1].replace("]", ""))
                production_rule_dict[parent].append((child, possibility))
            except Exception as e:
                print(e)
                print(grammar_line)
                raise RuntimeError(e)

        # sort the production according to the probability
        # sorted_production_rule_dict = dict(sorted(production_rule_dict.items(), key=lambda x: x[1], reverse=True))
        # print(sorted_production_rule_dict)
        IOUtils.dump(self.logging_file, [html_list_index_to_sheet_index], IOUtils.Format.txtList, append=True)

        row1_to_label = {}
        row1_to_noun_label = collections.defaultdict(str)
        row1_to_postag = {}

        each_html_len_index = 0
        pcfg_label_list = []
        pcfg_noun_label_list = []
        str_tree_list = []
        index_with_verb = []
        index_without_verb = []
        has_verb = False
        verb_count = 0
        total_count = 0
        for index, str_tree in enumerate(pos_tagged_html_list):
            parented_tree = ParentedTree.fromstring(str_tree)
            np_trees = self.traverse(parented_tree, production_rule_dict, "V")
            print("parented_tree", parented_tree)
            print("np_tree", np_trees)
            noun_np_trees = self.traverse(parented_tree, production_rule_dict, "N")
            print("parented_tree", parented_tree)
            print("noun_np_tree", noun_np_trees)
            if np_trees:
                try:
                    sorted_np_trees = sorted(np_trees, key=lambda x: x[1][0], reverse=True)
                    pcfg_label = " ".join(sorted_np_trees[0][0].leaves())
                    pcfg_label_list.append((pcfg_label, sorted_np_trees[0][1][1], sorted_np_trees[0][1][0]))
                    str_tree_list.append(str_tree)
                    index_with_verb.append(index)
                    has_verb = True
                except TypeError as e:
                    str_tree_list.append(str_tree)
                    IOUtils.dump(self.logging_file, [f"{e}"], IOUtils.Format.txtList, append=True)
                    IOUtils.dump(self.logging_file, [t for t in np_trees], IOUtils.Format.txtList, append=True)
                    index_without_verb.append(index)
                    raise RuntimeError(e)
            else:
                # if len(list(dict.fromkeys(parented_tree.leaves()))) == 1:
                #     pcfg_label = " ".join(list(dict.fromkeys(parented_tree.leaves())))
                # else:
                pcfg_label = ""
                pcfg_label_list.append((pcfg_label, 0, ""))
                str_tree_list.append(str_tree)
                index_without_verb.append(index)

            if noun_np_trees:
                try:
                    sorted_noun_np_trees = sorted(noun_np_trees, key=lambda x: x[1][0], reverse=True)
                    pcfg_noun_label = " ".join(sorted_noun_np_trees[0][0].leaves())
                    pcfg_noun_label_list.append((pcfg_noun_label, sorted_noun_np_trees[0][1][1], sorted_noun_np_trees[0][1][0]))
                except Exception as e:
                    raise RuntimeError(e)
            else:
                pcfg_noun_label_list.append(("", 0, ""))

            if index == each_html_len[each_html_len_index] - 1:
                total_count += 1
                if has_verb:
                    verb_count += 1
                has_verb = False
                pcfg_label_list = sorted(pcfg_label_list, key=lambda x: x[1], reverse=True)
                pcfg_noun_label_list = sorted(pcfg_noun_label_list, key=lambda x: x[1], reverse=True)
                if len(pcfg_label_list) == 0:
                    print(f"{has_verb=} {pcfg_label_list=}")
                row1_to_label[html_list_index_to_sheet_index[each_html_len_index]] = pcfg_label_list[0][0]
                if "VP" not in pcfg_label_list[0][2] and len(pcfg_noun_label_list) > 0:
                    row1_to_noun_label[html_list_index_to_sheet_index[each_html_len_index]] = pcfg_noun_label_list[0][0]
                # remove_non_english_words(lemmatization(tokenize(" ".join(pcfg_label_list_remove_duplicate))))
                row1_to_postag[html_list_index_to_sheet_index[each_html_len_index]] = "\n".join(str_tree_list)
                each_html_len_index += 1
                pcfg_label_list = []
                pcfg_noun_label_list = []
                str_tree_list = []
            if 0 < each_html_len_index < len(each_html_len) and each_html_len[each_html_len_index] == each_html_len[each_html_len_index - 1]:
                each_html_len_index += 1
                # print(html_list_index_to_sheet_index[each_html_len_index])
        # IOUtils.dump(self.logging_file, [index_with_verb, len(index_with_verb)], IOUtils.Format.txtList, append=True)
        # IOUtils.dump(self.logging_file, [index_without_verb, len(index_without_verb)], IOUtils.Format.txtList, append=True)
        # IOUtils.dump(self.logging_file, [verb_count, total_count, verb_count / total_count], IOUtils.Format.txtList, append=True)
        pcfg_wb_obj = openpyxl.Workbook()
        pcfg_sheet = pcfg_wb_obj.active
        IOUtils.dump(self.logging_file, [len(row1_to_label), row1_to_label], IOUtils.Format.txtList, append=True)
        pcfg_row = list(col_name_list)
        pcfg_row.append("pcfg label")
        pcfg_row.append("tree")
        pcfg_sheet.append(pcfg_row)
        for index, row in enumerate(load_xlsx(xlsx_path)):
            if row[index_of_manual_label] and row[index_of_manual_label] == "none":
                num_elements_not_visible += 1
                continue
            if index in row1_to_label:
                pcfg_label = row1_to_label[index]
                postag_tree = row1_to_postag[index]
            elif row[index_of_element_details] in element_details_to_sheet_index.keys():
                pcfg_label = row1_to_label[element_details_to_sheet_index[row[index_of_element_details]]]
                postag_tree = row1_to_postag[element_details_to_sheet_index[row[index_of_element_details]]]
            else:
                pcfg_label = ""
                postag_tree = ""
                IOUtils.dump(self.logging_file, [f"{row[index_of_id]} does not have label"], IOUtils.Format.txtList, append=True)
            if pcfg_label == "" or len(pcfg_label.split(" ")) >= 5:
                print("pcfg verb label does not exist", index, row[index_of_current_dom])
                pcfg_label = self.heuristic_rule_verb(row[index_of_current_dom], row[index_of_context_dom])
                print("pcfg label", pcfg_label)

            noun_label = ""
            if index in row1_to_noun_label:
                noun_label = row1_to_noun_label[index]
                print("row1_to_noun_label", row1_to_noun_label[index])
            elif row[index_of_element_details] in element_details_to_sheet_index.keys():
                noun_label = row1_to_noun_label[element_details_to_sheet_index[row[index_of_element_details]]]
                print("row1_to_noun_label", row1_to_noun_label[element_details_to_sheet_index[row[index_of_element_details]]])
            if len(noun_label.split(" ")) < 5:
                pcfg_label += " " + noun_label

            pcfg_row = list(row)
            remove_duplicate_word_pcfg_label = ""
            for word in pcfg_label.split(" "):
                if not word:
                    continue
                if word not in remove_duplicate_word_pcfg_label:
                    # remove_duplicate_word_pcfg_label = word + " " + remove_duplicate_word_pcfg_label
                    remove_duplicate_word_pcfg_label += " " + word
            pcfg_row.append(remove_duplicate_word_pcfg_label.strip())
            pcfg_row.append(postag_tree)
            pcfg_sheet.append(pcfg_row)

        pcfg_xlsx_path = f"{Macros.results_dir}/pcfg/{self.project_name}-pcfg-labels1.xlsx"
        pcfg_wb_obj.save(pcfg_xlsx_path)

    def heuristic_rule_verb(self, current_dom: str, context_dom: str):
        splitted_html = dom_dict_value_to_list(parse_dom_to_dict(current_dom))
        action_verb = []
        preprocessed = preprocess_dom(" ".join(splitted_html))
        for word in preprocessed.split(" "):
            word = word.strip()
            if verbnet.classids(word) and word not in action_verb:
                action_verb.append(word)
        if len(action_verb) > 0:
            return remove_non_english_words(" ".join(action_verb))
        for w in ["log in", "login", "Log in", "Login"]:
            if w in current_dom:
                return "log in"
        for w in ["log out", "logout", "Log out", "Logout"]:
            if w in current_dom:
                return "log out"
        for w in ["<ul>", "<li>", "<menu>", "<nav>", "<nav"]:
            if w in context_dom:
                return "navigate to"
        for w in ["<select>", "option"]: # role="option" onmousedown
            if w in current_dom:
                return "select"
        for w in ["OK", "ok", "Ok"]:
            if w in current_dom:
                return "confirm"
        return "view"

    def traverse(self, t: ParentedTree, production_rule_dict: dict, part_of_speech: str):
        np_trees = []
        try:
            t.label()
        except AttributeError:
            return np_trees
        if t.label().startswith(part_of_speech):
            current = t
            parent_label = current.parent().label()
            current_tree_position = current.treeposition()[-1]
            print(f"{t=} {parent_label=} {current_tree_position=}")
            try:
                possibility = production_rule_dict[parent_label][current_tree_position]
                IOUtils.dump(self.logging_file, [f"possibility {possibility}"], IOUtils.Format.txtList, append=True)
                np_trees.append([current, possibility])
            except IndexError as e:
                IOUtils.dump(self.logging_file, [f"{e} {t} {current.parent()} {current_tree_position}"], IOUtils.Format.txtList, append=True)
        for child in t:
            np_trees.extend(self.traverse(child, production_rule_dict, part_of_speech))
        return np_trees

    def pos_tag(self, html_list: List[str]):
        # 'parse', 'depparse', 'coref
        pos_tag_list = []
        with CoreNLPClient(
                annotators=['tokenize', 'ssplit', 'mwt', 'pos', 'parse'],
                timeout=30000,
                memory='6G') as client:
            for html in html_list:
                print("html", html)
                ann = client.annotate(html, output_format="json")
                tree_str = ann["sentences"][0]["parse"]
                pos_tag_list.append(tree_str)
        return pos_tag_list

    # for i in 'addressbook' 'claroline' 'collabtive' 'jpetstore' 'mantisbt' 'mrbs' 'phoenix' 'ppma';
    # do python -m ui.labels.PCFG --project_name $i pcfg_with_keybert; done;
    def pcfg_with_keybert(self):
        xlsx_path = f"{Macros.results_dir}/pcfg/{self.project_name}-pcfg-labels1.xlsx"

        col_name_list = column_name_list(xlsx_path)
        index_of_manual_label = col_name_list.index("manual label")
        index_of_pcfg_label = col_name_list.index("pcfg label")
        index_of_current_dom = col_name_list.index("current dom")
        index_of_context_dom = col_name_list.index("context dom")

        pcfg_wb_obj = openpyxl.Workbook()
        pcfg_sheet = pcfg_wb_obj.active
        pcfg_row = list(col_name_list)
        pcfg_row.append("pcfg-keybert label")
        pcfg_sheet.append(pcfg_row)

        for index, row in enumerate(load_xlsx(xlsx_path)):
            if row[index_of_manual_label] and row[index_of_manual_label] == "none":
                continue
            pcfg_label = row[index_of_pcfg_label]

            pcfg_row = list(row)
            keybert_edge_label = get_context_label(row[index_of_current_dom] + " " + pcfg_label)
            if keybert_edge_label == "":
                keybert_edge_label = get_context_label(row[index_of_context_dom] + " " + pcfg_label)

            pcfg_row.append(keybert_edge_label.strip())
            pcfg_sheet.append(pcfg_row)

        pcfg_xlsx_folder = Macros.results_dir / "pcfg-keybert"
        if not pcfg_xlsx_folder.exists():
            pcfg_xlsx_folder.mkdir()
        pcfg_xlsx_path = f"{Macros.results_dir}/pcfg-keybert/{self.project_name}-labels.xlsx"
        pcfg_wb_obj.save(pcfg_xlsx_path)


if __name__ == "__main__":
    CLI(PCFG, as_positional=False)

def dom_dict_to_pcfg_tree_dfs(input: Union[str, dict]) -> List[str]:
    if isinstance(input, str):
        # return ["(", input.strip(), ")"]
        return ["!" + input.strip()]

    if "tag" in input:
        toks = ["(", input["tag"]]
        for k, v in input.items():
            if k == "tag":
                continue
            elif k == "attributes":
                for attr_k, attr_v in v.items():
                    toks += ["(", attr_k] + dom_dict_to_pcfg_tree_dfs(attr_v) + [")"]
            else:
                toks += ["(", k] + dom_dict_to_pcfg_tree_dfs(v) + [")"]
        toks += [")"]
        return toks
    else:
        toks = ["(", "dict"]
        for k, v in input.items():
            toks += ["(", k] + dom_dict_to_pcfg_tree_dfs(v) + [")"]
        toks += [")"]
        return toks

def rank_pcfg_attributes(project_name):
    filename = project_name+"-pcfg.xlsx"
    freq_dict = collections.Counter()
    excel_path = Macros.results_dir / "pcfg" / f"{filename}"
    wb_obj = openpyxl.load_workbook(excel_path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    root_types = set()
    for i in range(1, m_row + 1):
        parent_obj = sheet_obj.cell(row=i, column=1)
        parent_value = parent_obj.value
        if parent_value == "anyaccess" or parent_value == "evlist":
            # these two attributes are added by crawljax
            continue
        root_types.add(parent_value)
        children_obj = sheet_obj.cell(row=i, column=2)
        children_value = children_obj.value
        percent_obj = sheet_obj.cell(row=i, column=3)
        percent_value = percent_obj.value
        if not children_value:
            continue
        freq_dict[parent_value] += 1
    freq_dict = {k: v for k, v in sorted(freq_dict.items(), key=lambda item: -item[1])}
    attribute_array = []
    freq_array = []
    for attribute, freq in freq_dict.items():
        attribute_array.append(attribute)
        freq_array.append(freq)
    # print("project:", project_name)
    # print(attribute_array)
    # print(freq_array)
    return attribute_array, freq_array


def pos_tag_with_pattern(html: str, pattern: str="VB"):
    with CoreNLPClient(
            annotators=['tokenize', 'ssplit', 'mwt', 'pos', 'parse'],
            timeout=30000,
            memory='6G') as client:
        # ann = client.annotate(tokenize(html))
        matches = client.tregex(html, pattern)
        print(matches)
        return " ".join(
            [sentence[match_id]['spanString'] for sentence in matches['sentences'] for match_id in sentence])


def parse_sentence_with_induced_grammar(grammar, tokens):
    parser = pchart.InsideChartParser(grammar)
    # parser = ViterbiParser(grammar)
    parser.trace(0)
    for p in parser.parse_all(tokens):
        print(p)


def phrase_extraction(project_name):
    result_path = Macros.results_dir / "pcfg" / f"{project_name}-phrase.log"
    IOUtils.rm(result_path)

    attribute_array, _ = rank_pcfg_attributes(project_name)
    IOUtils.dump(result_path, [attribute_array], IOUtils.Format.txtList, append=True)

    xlsx_path = f"{Macros.results_dir}/paths/{project_name}-labels.xlsx"
    wb_obj = openpyxl.load_workbook(xlsx_path)
    sheet_obj = wb_obj.active
    for index, row in enumerate(sheet_obj.iter_rows()):
        if index == 0:
            continue
        if row[2].value:
            verb = pos_tag_with_pattern(row[2].value, "VP|VB|VBD|VBG|VBN|VBP|VBZ")

            html_dict = json.loads(row[1].value)

            phrase = ""
            for attribute in attribute_array:
                if (attribute == "text") and (attribute in html_dict) and html_dict[attribute].strip() != "":
                    phrase = html_dict[attribute]
                    break
                if ("attributes" in html_dict) and (attribute in html_dict["attributes"]):
                    phrase = html_dict["attributes"][attribute]
                    break

            IOUtils.dump(result_path, [row[2].value.strip(), row[11].value, verb, phrase, "\n"], IOUtils.Format.txtList, append=True)